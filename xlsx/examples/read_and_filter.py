#!/usr/bin/env python3
"""
Example: Read an Excel file, filter rows based on criteria, write filtered results.
"""

from openpyxl import load_workbook, Workbook

def filter_excel(
    input_file: str,
    output_file: str,
    filter_column: int = 1,
    filter_value: str = None,
    min_value: float = None
):
    """
    Filter rows from an Excel file.
    
    Args:
        input_file: Source Excel file
        output_file: Destination for filtered data
        filter_column: 1-indexed column to filter on
        filter_value: Keep rows where column equals this value
        min_value: Keep rows where column value >= this number
    """
    wb_in = load_workbook(input_file)
    ws_in = wb_in.active
    
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Filtered"
    
    # Copy headers (row 1)
    for col, cell in enumerate(ws_in[1], 1):
        ws_out.cell(row=1, column=col, value=cell.value)
    
    # Filter and copy data rows
    out_row = 2
    for row in ws_in.iter_rows(min_row=2):
        cell_value = row[filter_column - 1].value
        
        keep = False
        if filter_value is not None and cell_value == filter_value:
            keep = True
        elif min_value is not None and isinstance(cell_value, (int, float)):
            keep = cell_value >= min_value
        
        if keep:
            for col, cell in enumerate(row, 1):
                ws_out.cell(row=out_row, column=col, value=cell.value)
            out_row += 1
    
    wb_out.save(output_file)
    print(f"Filtered {out_row - 2} rows → {output_file}")

# Example usage:
if __name__ == "__main__":
    # Filter sales_report.xlsx to keep only rows where column 5 (Total) >= 5000
    # filter_excel("sales_report.xlsx", "high_performers.xlsx", filter_column=5, min_value=5000)
    print("Usage: filter_excel(input, output, filter_column=N, filter_value='X' or min_value=N)")
