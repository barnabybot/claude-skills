#!/usr/bin/env python3
"""
Example: Create a formatted sales report with headers, data, and totals.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_sales_report(filename: str = "sales_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 Sales"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Product", "Jan", "Feb", "Mar", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    data = [
        ["Widget A", 1500, 1750, 2000],
        ["Widget B", 800, 950, 1100],
        ["Widget C", 2200, 2100, 2400],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
        
        # Total formula
        total_cell = ws.cell(row=row_idx, column=5)
        total_cell.value = f"=SUM(B{row_idx}:D{row_idx})"
        total_cell.border = border
        total_cell.font = Font(bold=True)
    
    # Column totals
    total_row = len(data) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 6):
        cell = ws.cell(row=total_row, column=col)
        col_letter = get_column_letter(col)
        cell.value = f"=SUM({col_letter}2:{col_letter}{total_row-1})"
        cell.font = Font(bold=True)
        cell.border = border
    
    # Auto-fit columns (approximate)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    wb.save(filename)
    print(f"Created: {filename}")

if __name__ == "__main__":
    create_sales_report()
